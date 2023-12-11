# flake8: noqa

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, date
import pymysql


pd.options.mode.chained_assignment = None


def pegarDados(
    dataInicio=0,
    dataFim=20230000,
    minCodColaborador=0,
    maxCodColaborador=99,
    nomeCliente="",
    fimNomeCliente="",
    atividade="",
    fimAtividade="",
    centroCusto="",
    fimCentroCusto="",
) -> tuple[tuple[tuple], tuple]:
    inicio = datetime.now()
    print(
        f"PARAMETROS: {dataInicio,dataFim,minCodColaborador,maxCodColaborador,nomeCliente, fimNomeCliente,atividade, fimAtividade, centroCusto, fimCentroCusto}"
    )
    if len(nomeCliente) == 0:
        nomeCliente = 0
    if len(atividade) == 0:
        atividade = 0
    if len(centroCusto) == 0:
        centroCusto = 0
    if len(fimNomeCliente) == 0:
        fimNomeCliente = 0
    if len(fimAtividade) == 0:
        fimAtividade = 0
    if len(fimCentroCusto) == 0:
        fimCentroCusto = "Z" * 10
    print(
        f"PARAMETROS: {dataInicio,dataFim,minCodColaborador,maxCodColaborador,nomeCliente, fimNomeCliente,atividade, fimAtividade, centroCusto, fimCentroCusto}"
    )
    connection = pymysql.connect(
        host="10.1.1.116", user="root", password="Ledzep0001", database="back_office"
    )
    with connection:
        with connection.cursor() as cursor:
            queryUpdate = """
            UPDATE
                ra_atividade_trabalho ra
                SET
                    ra.atividade_cod_atividade =   (SELECT
                cod_atividade
                    FROM
                        atividade
                    WHERE
                        desc_atividade = ra.desc_atividade limit 1)
                WHERE
                    atividade_cod_atividade is null"""
            cursor.execute(queryUpdate)
            connection.commit()

            querySelect = """
            SELECT
                this_.data_atividade_s as DataAtividade,
                this_.hora_inicio_s as HoraInicio,
                this_.hora_final_s as HoraFim,
                this_.atividade_cod_atividade as CodAtividade,
                this_.observacao1 as Observacao1,
                this_.observacao2 as Observacao2,
                pessoa.nome as nomeColaborador,
                ativ.desc_atividade as descricaoAtividade,
                em.nome_abrev as nomeCliente,
                ncc.descricao as CentroCusto
            FROM
                ra_atividade_trabalho this_
            INNER JOIN
                atividade ativ
                    on this_.atividade_cod_atividade=ativ.cod_atividade
            INNER JOIN
                colaboradores colab
                    ON this_.cod_colaborador=colab.cod_colaborador
            INNER JOIN
                pessoa
                    ON colab.cod_pessoa=pessoa.cod_pessoa
            INNER JOIN
                new_centro_custo ncc
                    ON ativ.centroCusto_id=ncc.id
            INNER JOIN
                emitente em
                    on this_.cod_emitente=em.cod_emitente
            WHERE
                this_.data_atividade_s>= %s
                AND this_.data_atividade_s<= %s
                AND this_.cod_colaborador>= %s
                AND this_.cod_colaborador<= %s
                AND em.nome_abrev>= %s
                AND em.nome_abrev<= %s
                AND ativ.desc_atividade>= %s
                AND ativ.desc_atividade<= %s
                AND ncc.codigo>= %s
                AND ncc.codigo<= %s;
                
            """
            dados = (
                dataInicio,
                dataFim,
                minCodColaborador,
                maxCodColaborador,
                nomeCliente,
                fimNomeCliente,
                atividade,
                fimAtividade,
                centroCusto,
                fimCentroCusto,
            )
            cursor.execute(querySelect, dados)

            ativTrabalho = cursor.fetchall()
            columns = (
                "Data",
                "HoraInicio",
                "HoraFim",
                "CodAtividade",
                "Observacao1",
                "Observacao2",
                "Colaborador",
                "DescricaoAtividade",
                "NomeCliente",
                "CentroCusto",
            )

    print(
        f"Tempo demorado para pegar {len(ativTrabalho)} registros do banco de dados: {inicio - datetime.now()}"
    )
    return ativTrabalho, columns  # type: ignore


def transformaParaDf(ativTrabalho, columns) -> pd.DataFrame:
    ativTrabalhoDf = pd.DataFrame(ativTrabalho, columns=columns)

    for cont, data in enumerate(ativTrabalhoDf["Data"]):
        if len(data) == 0:
            ativTrabalhoDf.iloc[cont, 0] = datetime(1970, 1, 1)
            continue
        ativTrabalhoDf.iloc[cont, 0] = datetime.strptime(data, "%Y%m%d")

    ativTrabalhoDf["Data"] = ativTrabalhoDf["Data"].sort_values()

    for cont, hora in enumerate(ativTrabalhoDf["HoraInicio"]):
        if len(hora) == 0 or hora == "null":
            ativTrabalhoDf.iloc[cont, 1] = datetime(1970, 1, 1)
            continue
        ativTrabalhoDf.iloc[cont, 1] = datetime(
            1970, 1, 1, hour=int(hora[:2]), minute=int(hora[2:4]), second=int(hora[4:])
        )

    for cont, hora in enumerate(ativTrabalhoDf["HoraFim"]):
        if len(hora) == 0 or hora == "null":
            ativTrabalhoDf.iloc[cont, 2] = datetime(1970, 1, 1)
            continue
        if int(hora[:2]) == 24:
            ativTrabalhoDf.iloc[cont, 2] = datetime(
                1970, 1, 1, hour=23, minute=59, second=59
            )
            continue
        ativTrabalhoDf.iloc[cont, 2] = datetime(
            1970, 1, 1, hour=int(hora[:2]), minute=int(hora[2:4]), second=int(hora[4:])
        )

    ativTrabalhoDf["Qtde Horas"] = (
        ativTrabalhoDf["HoraFim"] - ativTrabalhoDf["HoraInicio"]
    )

    ativTrabalhoDf["Qtde Horas"] = [
        str(a)[:4] for a in ativTrabalhoDf["Qtde Horas"].tolist()
    ]

    ativTrabalhoDf = ativTrabalhoDf.drop(columns=["HoraInicio", "HoraFim"])

    def temDescricao(row):
        if len(row["Observacao2"]) == 0:
            return f"{row['Observacao1']}"
        return f"{row['Observacao1']} + {row['Observacao2']}"

    try:
        ativTrabalhoDf["Descricao"] = ativTrabalhoDf.apply(
            lambda x: temDescricao(x), axis=1
        )
    except ValueError:
        return pd.DataFrame()

    ativTrabalhoDf = ativTrabalhoDf.drop(columns=["Observacao1", "Observacao2"])

    return ativTrabalhoDf


def gerarExcel(
    dataInicio: datetime.date,  # type: ignore
    dataFim: datetime.date,  # type: ignore
    minCodColaborador: int = 0,
    maxCodColaborador: int = 99,
    nomeCliente: str = "",
    fimNomeCliente: str = "",
    atividade: str = "",
    fimAtividade: str = "",
    centroCusto: str = "",
    fimCentroCusto: str = "",
) -> Workbook:
    ativTrabalho, columns = pegarDados(
        date.strftime(dataInicio, "%Y%m%d"),  # type: ignore
        date.strftime(dataFim, "%Y%m%d"),  # type: ignore
        minCodColaborador,
        maxCodColaborador,
        nomeCliente,
        fimNomeCliente,
        atividade,
        fimAtividade,
        centroCusto,
        fimCentroCusto,
    )
    
    ativTrabalhoDf = transformaParaDf(ativTrabalho, columns)
    print(f"Quantidade de registros: {len(ativTrabalhoDf)}")
    wb = Workbook()

    resumoSheet = wb.create_sheet(title='Resumo Atividade')

    resumoSheet['A1'] = 'Atividade'
    resumoSheet['B1'] = 'Quantidade Horas'
    c=1
    if len(ativTrabalhoDf) == 0:
        return wb  

    horasTotalDecimal = 0
    for desc_atividade in ativTrabalhoDf["DescricaoAtividade"].unique().tolist():
        if pd.isnull(desc_atividade):
            continue

        atividadeTrabalho = ativTrabalhoDf[
            ativTrabalhoDf["DescricaoAtividade"] == desc_atividade
        ]
        atividadeTrabalho["Data"] = [
            date(x.year, x.month, x.day) for x in atividadeTrabalho["Data"]
        ]
        atividadeTrabalho = atividadeTrabalho.sort_values(by="Data")

        atividadeTrabalho["Data"] = [
            date.strftime(data, "%d/%m/%Y") for data in atividadeTrabalho["Data"]
        ]

        qtdHoras = 0
        qtdMinutos = 0
        for x in atividadeTrabalho["Qtde Horas"]:
            if len(x) == 4:
                qtdHoras += int(x[0])
                qtdMinutos += int(x[2:4])

        horasDosMinutos = qtdMinutos // 60
        qtdHoras += horasDosMinutos
        qtdMinutos -= horasDosMinutos * 60
        if len(str(qtdMinutos)) == 1:
            qtdMinutos = '00'
        title = desc_atividade.replace(":", "-")
        title = title.replace("/", "|")

        ws = wb.create_sheet(title=title[:30])

        ws["E1"] = "Cliente:"
        ws["F1"] = str(atividadeTrabalho.iloc[0, 4])

        ws["E2"] = "Centro de custo:"
        ws["F2"] = str(atividadeTrabalho.iloc[0, 5])

        ws["E3"] = "Atividade:"
        ws["F3"] = str(title)

        ws["E4"] = "Qtde Horas Totais:"
        ws["F4"] = f"{qtdHoras}:{qtdMinutos}"
        def passarParaDecimal(row):
            hora: int = 0
            minuto: float = 0.0
            if len(row) == 4:
                hora = int(row[0])
                minuto: float = int(row[2:]) / 60
            elif len(row) == 5:
                hora = int(row[:2])
                minuto = int(row[3:]) / 60
            return hora + minuto
        
        c+=1
        horaMinuto =  f"{qtdHoras}:{qtdMinutos}"   
        horaMinutoDecimal = passarParaDecimal(horaMinuto)
        resumoSheet[f'A{c}'] = desc_atividade
        resumoSheet[f'B{c}'] = horaMinuto
        resumoSheet[f'C{c}'] = horaMinutoDecimal
        horasTotalDecimal += horaMinutoDecimal
        
        atividadeTrabalho["Qtde Horas Decimal"] = atividadeTrabalho["Qtde Horas"]
        
        atividadeTrabalho["Qtde Horas Decimal"] = atividadeTrabalho[
            "Qtde Horas Decimal"
        ].apply(lambda x: passarParaDecimal(x))

        column_widths = [13, 10, 10, 88, 30, 13, 40]
        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = column_width

        atividadeTrabalho = atividadeTrabalho[
            ["Data", "Qtde Horas", "Qtde Horas Decimal", "Descricao", "Colaborador"]
        ]

        for r in dataframe_to_rows(atividadeTrabalho, index=False, header=True):
            if "nxp/nxbrw\x08ndi069.w" in r:
                continue
            ws.append(r)

    del wb["Sheet"]
    column_widthsResumo = [50, 10]
    for i, column_width in enumerate(column_widthsResumo, 1):
        resumoSheet.column_dimensions[get_column_letter(i)].width = column_width
    c+=1
    
    resumoSheet[f'A{c}'] = 'TOTAL DE HORAS'
    resumoSheet[f'C{c}'] = horasTotalDecimal
    return wb


# print(gerarExcel(date(2023, 11, 6), date(2023, 11, 11), 0, 99, "", "", "", "", "", ""))
