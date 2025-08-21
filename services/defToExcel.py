import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, date
from services.selectData import selectDataFromDatabase
from services.utils import passarParaDecimal

pd.options.mode.chained_assignment = None


def transformaParaDf(ativTrabalho, columns) -> pd.DataFrame:
    ativTrabalhoDf = pd.DataFrame(ativTrabalho, columns=columns)

    for cont, data in enumerate(ativTrabalhoDf["Data"]):
        if len(data) == 0:
            ativTrabalhoDf.iloc[cont, 0] = datetime(1970, 1, 1)
            continue
        ativTrabalhoDf.iloc[cont, 0] = datetime.strptime(data, "%Y%m%d")

    ativTrabalhoDf["Data"] = ativTrabalhoDf["Data"].sort_values()

    for cont, hora in enumerate(ativTrabalhoDf["HoraInicio"]):
        if len(hora) == 0 or hora == "null":
            ativTrabalhoDf.iloc[cont, 1] = datetime(1970, 1, 1)
            continue
        ativTrabalhoDf.iloc[cont, 1] = datetime(
            1970, 1, 1, hour=int(hora[:2]), minute=int(hora[2:4]), second=int(hora[4:])
        )

    for cont, hora in enumerate(ativTrabalhoDf["HoraFim"]):
        if len(hora) == 0 or hora == "null":
            ativTrabalhoDf.iloc[cont, 2] = datetime(1970, 1, 1)
            continue
        if int(hora[:2]) == 24:
            ativTrabalhoDf.iloc[cont, 2] = datetime(
                1970, 1, 1, hour=23, minute=59, second=59
            )
            continue
        ativTrabalhoDf.iloc[cont, 2] = datetime(
            1970, 1, 1, hour=int(hora[:2]), minute=int(hora[2:4]), second=int(hora[4:])
        )

    ativTrabalhoDf["Qtde Horas"] = (
        ativTrabalhoDf["HoraFim"] - ativTrabalhoDf["HoraInicio"]
    )

    ativTrabalhoDf["Qtde Horas"] = [
        str(a)[:4] for a in ativTrabalhoDf["Qtde Horas"].tolist()
    ]

    ativTrabalhoDf = ativTrabalhoDf.drop(columns=["HoraInicio", "HoraFim"])

    def temDescricao(row):
        if len(row["Observacao2"]) == 0:
            return f"{row['Observacao1']}"
        return f"{row['Observacao1']} + {row['Observacao2']}"

    try:
        ativTrabalhoDf["Descricao"] = ativTrabalhoDf.apply(
            lambda x: temDescricao(x), axis=1
        )
    except ValueError:
        return pd.DataFrame()

    ativTrabalhoDf = ativTrabalhoDf.drop(columns=["Observacao1", "Observacao2"])

    return ativTrabalhoDf


def gerarExcel(
    dataInicio: str = "20230101",
    dataFim: str = "20230101",
    minCodColaborador: int | None = 0,
    maxCodColaborador: int | None = 99,
    nomeCliente: int | None = 0,
    fimNomeCliente: int | None = 0,
    atividade: int | None = 0,
    fimAtividade: int | None = 0,
    centroCusto: int | None = 0,
    fimCentroCusto: str | None = "",
) -> Workbook:
    ativTrabalho, columns = selectDataFromDatabase(
        dataInicio,
        dataFim,
        minCodColaborador,
        maxCodColaborador,
        nomeCliente,
        fimNomeCliente,
        atividade,
        fimAtividade,
        centroCusto,
        fimCentroCusto,
    )

    ativTrabalhoDf = transformaParaDf(ativTrabalho, columns)
    print(f"Quantidade de registros: {len(ativTrabalhoDf)}")
    wb = Workbook()

    resumoSheet = wb.create_sheet(title="Resumo Atividade")

    resumoSheet["A1"] = "Atividade"
    resumoSheet["B1"] = "Quantidade Horas"
    c = 1
    if len(ativTrabalhoDf) == 0:
        return wb

    horasTotalDecimal = 0
    for desc_atividade in ativTrabalhoDf["DescricaoAtividade"].unique().tolist():
        if pd.isnull(desc_atividade):
            continue

        atividadeTrabalho = ativTrabalhoDf[
            ativTrabalhoDf["DescricaoAtividade"] == desc_atividade
        ]
        atividadeTrabalho["Data"] = [
            date(x.year, x.month, x.day) for x in atividadeTrabalho["Data"]
        ]
        atividadeTrabalho = atividadeTrabalho.sort_values(by="Data")

        atividadeTrabalho["Data"] = [
            date.strftime(data, "%d/%m/%Y") for data in atividadeTrabalho["Data"]
        ]

        qtdHoras = 0
        qtdMinutos = 0
        for x in atividadeTrabalho["Qtde Horas"]:
            if len(x) == 4:
                qtdHoras += int(x[0])
                qtdMinutos += int(x[2:])
            elif len(x) == 5:
                qtdHoras += int(x[:2])
                qtdMinutos += int(x[3:])

        horasDosMinutos = qtdMinutos // 60
        qtdHoras += horasDosMinutos
        qtdMinutos -= horasDosMinutos * 60
        if len(str(qtdMinutos)) == 1:
            qtdMinutos = "00"
        title = desc_atividade.replace(":", "-")
        title = title.replace("/", "|")

        ws = wb.create_sheet(title=title[:30])

        ws["E1"] = "Cliente:"
        ws["F1"] = str(atividadeTrabalho.iloc[0, 4])

        ws["E2"] = "Centro de custo:"
        ws["F2"] = str(atividadeTrabalho.iloc[0, 5])

        ws["E3"] = "Atividade:"
        ws["F3"] = str(title)

        ws["E4"] = "Qtde Horas Totais:"
        ws["F4"] = f"{qtdHoras}:{qtdMinutos}"

        c += 1
        horaMinuto = f"{qtdHoras}:{qtdMinutos}"
        horaMinutoDecimal = passarParaDecimal(horaMinuto)
        resumoSheet[f"A{c}"] = desc_atividade
        resumoSheet[f"B{c}"] = horaMinuto
        resumoSheet[f"C{c}"] = horaMinutoDecimal
        horasTotalDecimal += horaMinutoDecimal

        atividadeTrabalho["Qtde Horas Decimal"] = atividadeTrabalho["Qtde Horas"]

        atividadeTrabalho["Qtde Horas Decimal"] = atividadeTrabalho[
            "Qtde Horas Decimal"
        ].apply(lambda x: passarParaDecimal(x))

        column_widths = [13, 10, 10, 88, 30, 13, 40]
        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = column_width

        atividadeTrabalho = atividadeTrabalho[
            ["Data", "Qtde Horas", "Qtde Horas Decimal", "Descricao", "Colaborador"]
        ]

        for r in dataframe_to_rows(atividadeTrabalho, index=False, header=True):
            if "nxp/nxbrw\x08ndi069.w" in r:
                continue
            ws.append(r)

    del wb["Sheet"]
    column_widthsResumo = [50, 10]
    for i, column_width in enumerate(column_widthsResumo, 1):
        resumoSheet.column_dimensions[get_column_letter(i)].width = column_width
    c += 1

    resumoSheet[f"A{c}"] = "TOTAL DE HORAS"
    resumoSheet[f"C{c}"] = horasTotalDecimal
    return wb
